from flask import Flask, render_template, request, send_file
import psycopg2
from psycopg2.extras import RealDictCursor
from datetime import datetime
import pandas as pd
import os
from math import radians, cos, sin, asin, sqrt
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage

app = Flask(__name__)

# CONFIGURATION
# Coordinates for 85 Grayston Drive, Morningside, Sandton
CAMPUS_LAT = -26.099059
CAMPUS_LON = 28.0538272
MAX_DISTANCE_KM = 0.2  # 200 meter allowed radius
START_DATE = datetime(2026, 2, 13)

def calculate_distance(lat1, lon1):
    """Calculates the distance between the student and the campus."""
    lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, CAMPUS_LON, CAMPUS_LAT])
    dlon, dlat = lon2 - lon1, lat2 - lat1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    return 2 * asin(sqrt(a)) * 6371

def get_db_connection():
    """Connects to the PostgreSQL database on Render."""
    db_url = os.environ.get('DATABASE_URL')
    # Fix for Render's postgres:// prefix requirement
    if db_url and db_url.startswith("postgres://"):
        db_url = db_url.replace("postgres://", "postgresql://", 1)
    conn = psycopg2.connect(db_url)
    return conn

def init_db():
    """Initializes the permanent database table."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS check_ins (
            id SERIAL PRIMARY KEY,
            first_name TEXT NOT NULL,
            middle_name TEXT,
            last_name TEXT NOT NULL,
            id_passport TEXT NOT NULL,
            email TEXT NOT NULL,
            phone TEXT NOT NULL,
            ethnicity TEXT,
            gender TEXT,
            course TEXT,
            week_number INTEGER,
            check_in_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    cur.close()
    conn.close()
    print("Southern Labs Database Initialized.")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    user_lat = request.form.get('latitude')
    user_lon = request.form.get('longitude')

    if not user_lat or not user_lon:
        return "<h1>Location Error</h1><p>GPS access is required to verify attendance.</p>"

    # Geofencing check
    dist = calculate_distance(float(user_lat), float(user_lon))
    if dist > MAX_DISTANCE_KM:
        return f"<h1>Check-in Denied</h1><p>You must be at 85 Grayston Drive. You are currently {round(dist*1000)}m away.</p>"

    today = datetime.now()
    week_num = max(1, ((today - START_DATE).days // 7) + 1)

    data = (
        request.form.get('first_name'), request.form.get('middle_name'),
        request.form.get('last_name'), request.form.get('id_passport'),
        request.form.get('email'), request.form.get('phone'),
        request.form.get('ethnicity'), request.form.get('gender'),
        request.form.get('course'), week_num
    )

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''INSERT INTO check_ins
        (first_name, middle_name, last_name, id_passport, email, phone, ethnicity, gender, course, week_number)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)''', data)
    conn.commit()
    cur.close()
    conn.close()
    return render_template('success.html', name=data[0])

@app.route('/admin')
def admin():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute('SELECT * FROM check_ins ORDER BY check_in_time DESC')
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('admin.html', rows=rows)

@app.route('/download/<string:course>/<int:week>')
def download(course, week):
    db_url = os.environ.get('DATABASE_URL')
    if db_url and db_url.startswith("postgres://"):
        db_url = db_url.replace("postgres://", "postgresql://", 1)
    
    # Filtering by week and course for separate spreadsheets
    df = pd.read_sql_query("SELECT * FROM check_ins WHERE week_number = %s AND course = %s",
                           db_url, params=(week, course))

    if df.empty:
        return f"<h1>No Data</h1><p>No check-ins for {course} in Week {week}.</p>"

    df.columns = ['ID', 'First Name', 'Middle Name', 'Last Name', 'ID/Passport', 'Email', 'Phone', 'Ethnicity', 'Gender', 'Course', 'Week', 'Time']

    os.makedirs('exports', exist_ok=True)
    filename = f"exports/SLABS_{course}_Week_{week}.xlsx"

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Attendance', startrow=5)
        ws = writer.sheets['Attendance']

        # 1. Insert Logo
        try:
            logo_path = os.path.join('static', 'images', 'logo.jpg')
            img = OpenpyxlImage(logo_path)
            img.width, img.height = 100, 100
            ws.add_image(img, 'A1')
        except:
            pass

        # 2. Header Formatting
        ws['D2'] = "Southern Labs Institute of Technology"
        ws['D2'].font = Font(size=16, bold=True, color='003366')
        ws['D3'] = f"Attendance Report: {course} - Week {week}"

        # 3. Table Styling
        fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for cell in ws[6]:
            cell.fill, cell.font = fill, header_font

        # 4. Signature line
        last_row = len(df) + 9
        ws.cell(row=last_row, column=1).value = "Facilitator Signature: ___________________________"
        ws.cell(row=last_row, column=1).font = Font(bold=True)

    return send_file(filename, as_attachment=True)

if __name__ == '__main__':
    # Initialize table before running
    try:
        init_db()
    except Exception as e:
        print(f"DB Error: {e}")
    
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)
